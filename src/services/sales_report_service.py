from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

WEEKDAY_KR = ["월", "화", "수", "목", "금", "토", "일"]


def format_currency(value: int | float) -> str:
    return f"{int(value):,}원"


def build_report_header(target_date: datetime) -> str:
    weekday = WEEKDAY_KR[target_date.weekday()]
    return f"{target_date.month}.{target_date.day} {weekday}"


def read_sales_values(
    excel_path: str | Path,
    sheet_name: str = "데일리매출",
    cash_cell: str = "M5",
    card_cell: str = "M6",
    transfer_cell: str = "M7",
    total_cell: str = "M8",
) -> dict[str, int]:
    workbook = load_workbook(excel_path, data_only=True)

    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"시트를 찾을 수 없습니다: {sheet_name}")

    sheet = workbook[sheet_name]

    values = {
        "cash": sheet[cash_cell].value or 0,
        "card": sheet[card_cell].value or 0,
        "transfer": sheet[transfer_cell].value or 0,
        "total": sheet[total_cell].value or 0,
    }

    return {key: int(value) for key, value in values.items()}


def build_sales_report_text(
    report_date: datetime,
    sales: dict[str, int],
) -> str:
    header = build_report_header(report_date)

    return (
        f"{header}\n\n"
        f"현금:\n{format_currency(sales['cash'])}\n\n"
        f"카드:\n{format_currency(sales['card'])}\n\n"
        f"계좌:\n{format_currency(sales['transfer'])}\n\n"
        f"총합:\n{format_currency(sales['total'])}"
    )


# ── 월 목표 카운트다운 ────────────────────────────────────────────────

def read_daily_section_totals(daily_path: str | Path) -> dict[str, int]:
    """데일리 파일에서 센터/레슨 매출 합계를 반환한다."""
    from src.services.entry_reader_service import read_sales_entries
    entries = read_sales_entries(daily_path)
    return {
        "center": sum(e.amount for e in entries if e.section == "센터"),
        "pt": sum(e.amount for e in entries if e.section == "레슨"),
    }



# 총매출 파일 월 시트에서 센터/피티 금액 컬럼 (1-based)
# 센터: 시작 B(2) → 이름 D(4), 금액 G(7)
# 레슨: 시작 P(16) → 이름 R(18), 금액 U(21)
_MONTHLY_SECTION_COLS = {
    "center": (4, 7),
    "pt": (18, 21),
}
_MONTHLY_DATA_START_ROW = 12


def read_monthly_totals_by_section(
    total_sales_path: str | Path,
    year: int,
    month: int,
    password: str | None = None,
) -> dict[str, int]:
    """총매출 파일의 해당 월 시트에서 센터/피티 매출 합계를 반환한다."""
    from src.services.total_sales_service import find_monthly_sheet_name, open_workbook

    wb = open_workbook(total_sales_path, password=password)
    try:
        sheet_name = find_monthly_sheet_name(wb.sheetnames, year, month)
        ws = wb[sheet_name]
        totals: dict[str, int] = {"center": 0, "pt": 0}
        max_row = ws.max_row or 500
        for key, (name_col, amount_col) in _MONTHLY_SECTION_COLS.items():
            for row_num in range(_MONTHLY_DATA_START_ROW, max_row + 1):
                name_val = ws.cell(row=row_num, column=name_col).value
                if not name_val or (isinstance(name_val, str) and name_val.startswith("=")):
                    continue
                amount_val = ws.cell(row=row_num, column=amount_col).value
                if isinstance(amount_val, (int, float)) and amount_val > 0:
                    totals[key] += int(amount_val)
        return totals
    finally:
        wb.close()


def build_countdown_text(
    today_center: int,
    today_pt: int,
    center_target: int,
    pt_target: int,
    running_center: int,
    running_pt: int,
) -> str:
    today_total = today_center + today_pt

    def _remaining(target: int, running: int) -> str:
        diff = target - running
        return "달성! 🎉" if diff <= 0 else format_currency(diff)

    return "\n".join([
        f"센터: {format_currency(today_center)}",
        f"피티: {format_currency(today_pt)}",
        f"총합: {format_currency(today_total)}",
        "",
        "목표까지 앞으로",
        f"센터: {_remaining(center_target, running_center)}",
        f"피티: {_remaining(pt_target, running_pt)}",
    ])
