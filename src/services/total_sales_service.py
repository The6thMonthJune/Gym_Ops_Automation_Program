from __future__ import annotations

import io
from pathlib import Path

import openpyxl


def find_monthly_sheet_name(sheet_names: list[str], year: int, month: int) -> str:
    """
    시트 이름 목록에서 연도/월에 해당하는 매출 시트명을 찾아 반환한다.
    openpyxl Workbook과 xlwings Book 모두에서 호출할 수 있도록
    시트 이름 목록(list)을 인자로 받는다.

    허용 패턴 예시:
      '총 매출25년_9월', '총 매출26년 4월', '총 매출26년 03월'
    """
    two_digit_year = year % 100
    year_token = f"{two_digit_year}년"
    month_tokens = {f"{month}월", f"{month:02d}월"}

    for name in sheet_names:
        if "매출" not in name:
            continue
        if year_token not in name:
            continue
        for mt in month_tokens:
            if mt in name:
                return name

    raise ValueError(
        f"{year}년 {month}월에 해당하는 매출 시트를 찾을 수 없습니다.\n"
        f"파일 내 시트 목록: {', '.join(sheet_names)}"
    )


def find_monthly_sheet(wb, year: int, month: int) -> str:
    """openpyxl Workbook 전용 래퍼."""
    return find_monthly_sheet_name(wb.sheetnames, year, month)


def find_monthly_expense_sheet_name(sheet_names: list[str], year: int, month: int) -> str:
    """
    시트 이름 목록에서 연도/월에 해당하는 지출 시트명을 찾아 반환한다.

    허용 패턴 예시:
      '총 지출26년 4월', '총 지출26년 04월', '총 지출25년_9월'
    """
    two_digit_year = year % 100
    year_token = f"{two_digit_year}년"
    month_tokens = {f"{month}월", f"{month:02d}월"}

    for name in sheet_names:
        if "지출" not in name:
            continue
        if year_token not in name:
            continue
        for mt in month_tokens:
            if mt in name:
                return name

    raise ValueError(
        f"{year}년 {month}월에 해당하는 지출 시트를 찾을 수 없습니다.\n"
        f"파일 내 시트 목록: {', '.join(sheet_names)}\n"
        f"시트 이름 형식 예시: '총 지출{two_digit_year}년 {month}월'"
    )


def open_workbook(file_path: str | Path, password: str | None = None):
    """
    엑셀 파일을 openpyxl로 열어 반환한다. (읽기 전용 용도)
    password가 있으면 msoffcrypto-tool로 암호 해제 후 열어 반환한다.
    """
    if password:
        try:
            import msoffcrypto
        except ImportError:
            raise ImportError(
                "암호 해제를 위해 msoffcrypto-tool 패키지가 필요합니다.\n"
                "pip install msoffcrypto-tool"
            )
        with open(file_path, "rb") as f:
            office_file = msoffcrypto.OfficeFile(f)
            office_file.load_key(password=password)
            decrypted = io.BytesIO()
            office_file.decrypt(decrypted)
        return openpyxl.load_workbook(decrypted, data_only=True)

    return openpyxl.load_workbook(file_path, data_only=True)
