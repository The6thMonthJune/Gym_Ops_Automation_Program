from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import openpyxl

from src.config.constants import DEFAULT_SHEET_NAME
from src.config.settings import get_expense_daily_sheet

# 매출 섹션 시작 열
_SECTION_START_COL: dict[str, int] = {"센터": 2, "레슨": 16}
_SALES_START_ROW = 12

# 매출 컬럼 오프셋 (섹션 시작 기준)
# entry_service.py _write_entry_row 기준:
# +0=계약일, +1=일, +2=회원명, +3=종목, +4=회원권, +5=금액(부가세),
# +6=금액(부가제외/수식), +7=결제방법, +8=승인번호, +9=FC, +10=담당
_OFF_DAY = 1
_OFF_NAME = 2
_OFF_CATEGORY = 3
_OFF_MEMBERSHIP = 4
_OFF_AMOUNT = 5
_OFF_PAYMENT = 7
_OFF_APPROVAL = 8
_OFF_FC = 9
_OFF_MANAGER = 10

# 지출 컬럼 (expense_service와 동일)
_EXP_START_ROW = 6
_EXP_COL_DATE = 3
_EXP_COL_CAT = 4
_EXP_COL_DESC = 5
_EXP_COL_AMT = 6
_EXP_COL_PAY = 7
_EXP_COL_MGR = 8
_EXP_COL_VND = 9
_EXP_COL_NOTE = 10


@dataclass
class SalesEntryRow:
    row_num: int
    section: str
    day: int
    name: str
    category: str
    membership: str
    amount: int
    payment_method: str
    approval_number: str
    fc: str
    manager: str


@dataclass
class ExpenseEntryRow:
    row_num: int
    day: int
    category: str
    description: str
    amount: int
    payment_method: str
    manager: str
    vendor: str
    note: str


def _cell(row: tuple, col_1based: int):
    """values_only 튜플에서 1-based 컬럼 값을 안전하게 읽는다."""
    idx = col_1based - 1
    return row[idx] if idx < len(row) else None


def read_sales_entries(file_path: str | Path) -> list[SalesEntryRow]:
    """매출 내역을 읽는다.
    read_only=True 모드에서 ws.cell() 임의 접근은 스트리밍 XML을 매번 재탐색해
    행 수가 많으면 UI 스레드를 프리즈시킨다. iter_rows로 교체해 해결.
    """
    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    try:
        if DEFAULT_SHEET_NAME not in wb.sheetnames:
            return []
        ws = wb[DEFAULT_SHEET_NAME]
        entries: list[SalesEntryRow] = []

        max_col = 28  # AA열(레슨 마지막 열 16+11=27) 커버
        rows = list(ws.iter_rows(min_row=_SALES_START_ROW, max_col=max_col, values_only=True))

        for section, col_start in _SECTION_START_COL.items():
            for i, row in enumerate(rows):
                name = _cell(row, col_start + _OFF_NAME)
                if name is None or (isinstance(name, str) and name.startswith("=")):
                    continue
                day_val    = _cell(row, col_start + _OFF_DAY)
                amount_val = _cell(row, col_start + _OFF_AMOUNT)
                # 조건부서식/유효성검사 더미 행 제거: 일(day)이 없거나 금액이 0인 경우 건너뜀
                if day_val is None or (amount_val is not None and amount_val == 0 and day_val == 0):
                    continue
                entries.append(SalesEntryRow(
                    row_num=_SALES_START_ROW + i,
                    section=section,
                    day=int(day_val) if day_val is not None else 0,
                    name=str(name).strip(),
                    category=str(_cell(row, col_start + _OFF_CATEGORY) or "").strip(),
                    membership=str(_cell(row, col_start + _OFF_MEMBERSHIP) or "").strip(),
                    amount=int(amount_val) if amount_val is not None else 0,
                    payment_method=str(_cell(row, col_start + _OFF_PAYMENT) or "").strip(),
                    approval_number=str(_cell(row, col_start + _OFF_APPROVAL) or "").strip(),
                    fc=str(_cell(row, col_start + _OFF_FC) or "").strip(),
                    manager=str(_cell(row, col_start + _OFF_MANAGER) or "").strip(),
                ))

        return sorted(entries, key=lambda e: (e.day, e.row_num))
    finally:
        wb.close()


def read_expense_entries(file_path: str | Path) -> list[ExpenseEntryRow]:
    sheet_name = get_expense_daily_sheet()
    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        entries: list[ExpenseEntryRow] = []

        rows = list(ws.iter_rows(min_row=_EXP_START_ROW, max_col=_EXP_COL_NOTE, values_only=True))
        for i, row in enumerate(rows):
            date_val = _cell(row, _EXP_COL_DATE)
            if date_val is None:
                continue
            if isinstance(date_val, (date, datetime)):
                day = date_val.day
            else:
                continue
            amount_val = _cell(row, _EXP_COL_AMT)
            entries.append(ExpenseEntryRow(
                row_num=_EXP_START_ROW + i,
                day=day,
                category=str(_cell(row, _EXP_COL_CAT) or "").strip(),
                description=str(_cell(row, _EXP_COL_DESC) or "").strip(),
                amount=int(amount_val) if amount_val is not None else 0,
                payment_method=str(_cell(row, _EXP_COL_PAY) or "").strip(),
                manager=str(_cell(row, _EXP_COL_MGR) or "").strip(),
                vendor=str(_cell(row, _EXP_COL_VND) or "").strip(),
                note=str(_cell(row, _EXP_COL_NOTE) or "").strip(),
            ))

        return entries
    finally:
        wb.close()


def calc_total_sales(entries: list[SalesEntryRow]) -> int:
    return sum(e.amount for e in entries)
