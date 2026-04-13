from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

from src.core.file_naming import build_next_date_path
from src.config.constants import DEFAULT_SHEET_NAME
from src.config.settings import get_expense_daily_sheet


def create_next_daily_file(source_file: str | Path, overwrite: bool = False) -> Path:
    """
    원본 파일을 복사하여 다음 날짜 파일을 생성한다.
    복사 후 매출·지출 입력 데이터를 초기화하고 수식·유효성 검사는 보존한다.
    """
    source_path = Path(source_file)

    if not source_path.exists():
        raise FileNotFoundError(f"원본 파일을 찾을 수 없습니다: {source_path}")

    target_path = build_next_date_path(source_path)

    if target_path.exists() and not overwrite:
        raise FileExistsError(f"이미 대상 파일이 존재합니다: {target_path}")

    shutil.copy2(source_path, target_path)
    _clear_daily_entries(target_path)
    return target_path


# ── 데이터 초기화 ──────────────────────────────────────────────────────────────

def _is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _clear_value(cell) -> None:
    """수식이 아닌 경우에만 셀 값을 지운다."""
    if cell.value is not None and not _is_formula(cell.value):
        cell.value = None


def _clear_daily_entries(file_path: Path) -> None:
    """복사된 데일리 파일의 매출·지출 데이터를 초기화한다."""
    wb = openpyxl.load_workbook(str(file_path))

    sales_name = DEFAULT_SHEET_NAME          # "데일리매출"
    expense_name = get_expense_daily_sheet() # "데일리지출" (설정값)

    if sales_name in wb.sheetnames:
        _clear_sales_sheet(wb[sales_name])

    if expense_name in wb.sheetnames:
        _clear_expense_sheet(wb[expense_name])

    wb.save(str(file_path))


def _clear_sales_sheet(ws) -> None:
    """데일리매출 시트 초기화 (12행~, 센터 B-M · 레슨 P-AA).
    H·M열의 VAT/자동기타 수식은 보존한다."""
    CENTER_COLS = range(2, 14)   # B–M
    LESSON_COLS = range(16, 28)  # P–AA (레슨 섹션)

    for row_num in range(12, min(ws.max_row + 1, 400)):
        center_name = ws.cell(row_num, 4).value   # D: 회원명 (센터)
        lesson_name = ws.cell(row_num, 18).value  # R: 회원명 (레슨, 16+2)

        has_center = center_name is not None and not _is_formula(center_name)
        has_lesson = lesson_name is not None and not _is_formula(lesson_name)

        if not has_center and not has_lesson:
            continue

        for col in CENTER_COLS:
            _clear_value(ws.cell(row_num, col))
        for col in LESSON_COLS:
            _clear_value(ws.cell(row_num, col))


def _clear_expense_sheet(ws) -> None:
    """데일리지출 시트 초기화 (6행~, A-J열).
    B열 미리 채워진 순번(C열이 None인 행)은 건드리지 않는다."""
    for row_num in range(6, ws.max_row + 1):
        date_val = ws.cell(row_num, 3).value  # C: 일자 (정상 입력)
        bug_val  = ws.cell(row_num, 1).value  # A: 구버전 버그로 잘못 기록된 값

        if date_val is None and bug_val is None:
            continue  # B열 순번만 있는 빈 행 → 보존

        for col in range(1, 11):  # A–J
            _clear_value(ws.cell(row_num, col))