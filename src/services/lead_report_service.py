from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.services.lead_service import CHANNELS, load_leads


# ── 색상 팔레트 ──────────────────────────────────────────────────────
_NAVY   = "1E2D3D"
_BLUE   = "2563EB"
_LBLUE  = "DBEAFE"
_GRAY   = "F3F4F6"
_WHITE  = "FFFFFF"
_BLACK  = "111827"


def _header_font(white: bool = True) -> Font:
    return Font(name="맑은 고딕", bold=True, color=_WHITE if white else _BLACK, size=10)


def _body_font() -> Font:
    return Font(name="맑은 고딕", size=10)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _thin_border() -> Border:
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def _apply_header_row(ws, row: int, values: list, fill_color: str = _NAVY) -> None:
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = _header_font()
        cell.fill = _fill(fill_color)
        cell.alignment = _center()
        cell.border = _thin_border()


def _apply_data_row(ws, row: int, values: list, shade: bool = False) -> None:
    bg = _GRAY if shade else _WHITE
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = _body_font()
        cell.fill = _fill(bg)
        cell.alignment = _center()
        cell.border = _thin_border()


# ── 시트 1: 채널별 현황 ───────────────────────────────────────────────

def _build_channel_sheet(ws, leads) -> None:
    ws.title = "채널별 현황"

    # 월 목록 추출 (최근 순)
    months: list[tuple[int, int]] = sorted(
        {(l.contract_date.year, l.contract_date.month) for l in leads},
        reverse=True,
    )

    # 헤더
    _apply_header_row(ws, 1, ["월"] + CHANNELS + ["합계"])

    # 데이터 행
    for r, (yr, mo) in enumerate(months, 2):
        month_leads = [l for l in leads if l.contract_date.year == yr and l.contract_date.month == mo]
        counts = {ch: 0 for ch in CHANNELS}
        for lead in month_leads:
            key = lead.channel if lead.channel in counts else "기타"
            counts[key] += 1
        row_vals = [f"{yr}년 {mo}월"] + [counts[ch] for ch in CHANNELS] + [sum(counts.values())]
        _apply_data_row(ws, r, row_vals, shade=(r % 2 == 0))

    # 합계 행
    if months:
        total_row = len(months) + 2
        grand = {ch: 0 for ch in CHANNELS}
        for lead in leads:
            key = lead.channel if lead.channel in grand else "기타"
            grand[key] += 1
        total_vals = ["합계"] + [grand[ch] for ch in CHANNELS] + [sum(grand.values())]
        for col, val in enumerate(total_vals, 1):
            cell = ws.cell(row=total_row, column=col, value=val)
            cell.font = _header_font(white=False)
            cell.fill = _fill(_LBLUE)
            cell.alignment = _center()
            cell.border = _thin_border()

    # 열 너비
    ws.column_dimensions["A"].width = 14
    for col in range(2, len(CHANNELS) + 3):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.row_dimensions[1].height = 20


# ── 시트 2: 거주지역별 현황 ──────────────────────────────────────────

def _build_district_sheet(ws, leads) -> None:
    ws.title = "거주지역별 현황"

    months: list[tuple[int, int]] = sorted(
        {(l.contract_date.year, l.contract_date.month) for l in leads},
        reverse=True,
    )
    all_districts: list[str] = sorted(
        {l.residence_district or "미입력" for l in leads}
    )

    _apply_header_row(ws, 1, ["월"] + all_districts + ["합계"])

    for r, (yr, mo) in enumerate(months, 2):
        month_leads = [l for l in leads if l.contract_date.year == yr and l.contract_date.month == mo]
        counts: dict[str, int] = {d: 0 for d in all_districts}
        for lead in month_leads:
            d = lead.residence_district or "미입력"
            counts[d] = counts.get(d, 0) + 1
        row_vals = [f"{yr}년 {mo}월"] + [counts[d] for d in all_districts] + [sum(counts.values())]
        _apply_data_row(ws, r, row_vals, shade=(r % 2 == 0))

    if months:
        total_row = len(months) + 2
        grand: dict[str, int] = {d: 0 for d in all_districts}
        for lead in leads:
            d = lead.residence_district or "미입력"
            grand[d] = grand.get(d, 0) + 1
        total_vals = ["합계"] + [grand[d] for d in all_districts] + [sum(grand.values())]
        for col, val in enumerate(total_vals, 1):
            cell = ws.cell(row=total_row, column=col, value=val)
            cell.font = _header_font(white=False)
            cell.fill = _fill(_LBLUE)
            cell.alignment = _center()
            cell.border = _thin_border()

    ws.column_dimensions["A"].width = 14
    for col in range(2, len(all_districts) + 3):
        ws.column_dimensions[get_column_letter(col)].width = max(10, len(all_districts[col - 2]) * 2 + 2 if col - 2 < len(all_districts) else 10)
    ws.row_dimensions[1].height = 20


# ── 시트 3: 원본 데이터 ──────────────────────────────────────────────

def _build_raw_sheet(ws, leads) -> None:
    ws.title = "원본 데이터"
    headers = ["계약일", "성함", "가입경로", "기타상세", "거주지역", "등록구분", "입력방식"]
    _apply_header_row(ws, 1, headers)
    for r, lead in enumerate(leads, 2):
        row_vals = [
            lead.contract_date.isoformat(),
            lead.member_name,
            lead.channel,
            lead.channel_detail or "",
            lead.residence_district or "",
            lead.registration_type,
            lead.source,
        ]
        _apply_data_row(ws, r, row_vals, shade=(r % 2 == 0))

    col_widths = [14, 12, 10, 14, 12, 10, 10]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20


# ── 공개 API ─────────────────────────────────────────────────────────

def generate_report(save_path: str | Path) -> Path:
    """
    전체 유입경로 데이터를 읽어 3시트짜리 엑셀 보고서를 생성한다.
    save_path: 저장할 파일 경로 (.xlsx)
    """
    leads = load_leads()
    if not leads:
        raise ValueError("저장된 유입경로 데이터가 없습니다.")

    wb = Workbook()
    ws_channel = wb.active
    _build_channel_sheet(ws_channel, leads)

    ws_district = wb.create_sheet()
    _build_district_sheet(ws_district, leads)

    ws_raw = wb.create_sheet()
    _build_raw_sheet(ws_raw, leads)

    out = Path(save_path)
    wb.save(str(out))
    return out
